from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
import pandas as pd

logger = logging.getLogger(__name__)

class ExcelChart():
    def __init__(
        self,
        chart_df: pd.DataFrame,
        raw_df_list: list,
        start_date: str, 
        end_date: str,
        filename: str, 
        filepath: str
    ):
        """Create a new excel sheet with a line chart as the first sheet, 
        append raw dataframes as separate worksheets, 
        then save the excel workbook.

        Args:
            chart_df (pd.DataFrame): pandas DataFrame used to plot line chart of indexed price history data
            raw_df_list (list): list of raw and indexed dataframes of price data for each financial asset
            start_date (str): start date of line chart
            end_date (str): end date of line chart
            filename (str): file name of excel file
            filepath (str): absolute file path to save excel file to
        """
        self.chart_df = chart_df
        self.raw_df_list = raw_df_list
        self.start_date = start_date
        self.end_date = end_date
        self.filename = filename
        self.filepath = filepath
    
    ###########
    # Helpers
    ###########

    def format_date_display(
        self,
        start_date: str,
        end_date: str):
        """Converts start date and end date from DD/MM/YYYY format to dd month_name yy format for excel chart display.

        Args:
            start_date (str): start date in string in DD/MM/YYYY format
            end_date (str):  end date in string in DD/MM/YYYY format

        Returns:
            start_date (str): start date in string in dd month_name yy format
            end_date (str):  end date in string in dd month_name yy format
        """

        start_date = datetime.strptime(start_date, '%d/%m/%Y') # read str as datetime
        start_date = datetime.strftime(start_date, '%d %b %y') # convert to dd month_name yy format

        end_date = datetime.strptime(end_date, '%d/%m/%Y') # read str as datetime
        end_date = datetime.strftime(end_date, '%d %b %y') # convert to dd month_name yy format

        return start_date, end_date

    def new_excelsheet(
        self,
        df: pd.DataFrame, 
        sheetname: str, 
        sheetnumber: int = 0
    ):
        """Writes a pandas dataframe as a new excel spreadsheet in an excel workbook

        Args:
            df (pd.DataFrame): pandas dataframe to be written as an excel spreadsheet
            sheetname (str): name of new excel spreadsheet
            sheetnumber (int, optional): sheet number of the new excel spreadsheet. 
            Defaults to 0 (default sheet/first displayed sheet)

        Returns:
            wb (object): excel Workbook object
        """
        # remove timestamp from datetime index
        df.index = df.index.date 

        wb = Workbook()
        newsheet = wb.create_sheet(sheetname, sheetnumber)
        
        for row in dataframe_to_rows(df, index = True, header = True):
            newsheet.append(row)
            
        for cell in newsheet[1]: # excel indexing starts from 1, to select row 1
            cell.style = 'Pandas'
        
        # Drop default blank 'Sheet' in workbook
        wb.remove(wb['Sheet'])
        
        return wb

    def append_excelsheet(
        self,
        wb: object, 
        df: pd.DataFrame, 
        sheetname: str, 
        sheetnumber: int =0
    ):
        """Appends a pandas dataframe as an additional excel spreadsheet to an existing excel Workbook

        Args:
            wb (object): excel Workbook object 
            df (pd.DataFrame): pandas dataframe to append as a new excel spreadsheet
            sheetname (str): name of newly appended worksheet
            sheetnumber (int, optional): Sheet number of newly appended worksheet in excel workbook. 
            Defaults to 0 (first displayed worksheet).

        Returns:
            wb (object): excel workbook object with appended worksheet
        """
        # remove timestamp from datetime index
        df.index = df.index.date 
        
        newsheet = wb.create_sheet(sheetname, sheetnumber)
        
        for row in dataframe_to_rows(df, index = True, header = True):
            newsheet.append(row)
            
        for cell in newsheet[1]: # excel indexing starts from 1, to select row 1
            cell.style = 'Pandas'
        
        return wb

    def create_linechart(
        self,
        wb: object,
        title: str,
        ytitle: str,
        date_unit: str,
        date_format: str,
        xtitle: str = "Date",
        chartcell: str ='G3'
    ):
        """Creates a new time-series line chart from data wthin an excel workbook object 
        using Date as the first column.

        Args:
            wb (object): excel workbook object
            title (str): chart title
            ytitle (str): y axis title
            date_unit (str): major date units for x axis in 'days' or 'months' format.
            date_format (str): format for displayed date.
            xtitle (str, optional): x axis title. Defaults to "Date".
            chartcell (str, optional): cell in workbook object to place chart. Defaults to 'G3'.

        Returns:
            wb (object): excel workbook object with newly created linechart
        """
        
        sheet = wb.active
        
        # exclude date column when selecting data, include header
        data = Reference(
            worksheet=sheet,
            min_row=1,
            max_row=sheet.max_row,
            min_col=2,
            max_col=sheet.max_column
        )

        # select only date column without title
        dates = Reference(
            worksheet=sheet,
            min_col=1,
            max_col=1,
            min_row=2, 
            max_row=sheet.max_row
        )
        
        chart = LineChart()
        chart.title = title
        chart.style = 2
        
        chart.y_axis.title = ytitle
        chart.x_axis.title = xtitle
        
        chart.y_axis.crossAx = 500
        chart.x_axis = DateAxis(crossAx=100)
        chart.x_axis.number_format = date_format
        chart.x_axis.majorTimeUnit = date_unit
        
        chart.height = 14
        chart.width = 25
        
        chart.add_data(data, titles_from_data=True)
        
        chart.set_categories(dates)
        
        sheet.add_chart(chart, chartcell)

        return wb
    
    #################
    # Core function
    #################

    def run(self):
        """Create a new excel sheet with line chart, append raw dataframes as separate worksheets, 
        then save the excel workbook.
        """
        
        excel = self.new_excelsheet(
            df = self.chart_df, 
            sheetname = 'Chart', 
            sheetnumber = 0
        )

        start_date, end_date = self.format_date_display(
            self.start_date,
            self.end_date
        )

        excel = self.create_linechart(
            wb = excel,
            title = f"Indexed price history ({start_date} - {end_date})",
            ytitle = f"Change in asset price relative to reference value on {start_date}",
            date_unit = "months",
            date_format = "mmm-yy"
        )

        for list_index, df in enumerate(self.raw_df_list):  

            excel = self.append_excelsheet(
                wb = excel, 
                df = df, 
                sheetname = df.name,
                sheetnumber = list_index + 1
            )

        excel.save(f'{self.filepath}/{self.filename}.xlsx')
        logger.info(f'Saved {self.filepath}/{self.filename}.xlsx')

        excel.close()